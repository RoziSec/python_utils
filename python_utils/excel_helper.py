# -*- coding: utf-8 -*-
"""Excel 操作类。

基于 ``openpyxl`` 读写 ``.xlsx``，以“字典列表 <-> 工作表”为核心用法，
适合把查询结果导出报表、或读取表格数据。

依赖：``pip install openpyxl``（延迟导入，用到才需要）。

    from python_utils import ExcelHelper

    rows = [{"姓名": "张三", "年龄": 20}, {"姓名": "李四", "年龄": 25}]
    ExcelHelper.write("out.xlsx", rows)                 # 导出
    data = ExcelHelper.read("out.xlsx")                 # 读取为字典列表
    ExcelHelper.write_sheets("multi.xlsx", {"用户": rows, "订单": [...]})
"""

from __future__ import annotations

from typing import Any, Dict, List, Optional, Union
from pathlib import Path

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None


def _require():
    if openpyxl is None:
        raise ImportError("使用 ExcelHelper 需要先安装：pip install openpyxl")


class ExcelHelper:
    """Excel 读写工具类（静态方法）。"""

    @staticmethod
    def read(path: Union[str, Path], sheet: Optional[Union[str, int]] = None,
             *, header: bool = True) -> List[Dict[str, Any]]:
        """读取工作表为字典列表（首行为表头）。

        :param sheet: 工作表名或索引；默认活动表。
        :param header: 首行是否为表头；False 则用列号 col0/col1... 作键。
        """
        _require()
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active if sheet is None else (wb[sheet] if isinstance(sheet, str) else wb.worksheets[sheet])
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        if header:
            keys = [str(c) if c is not None else f"col{i}" for i, c in enumerate(rows[0])]
            body = rows[1:]
        else:
            keys = [f"col{i}" for i in range(len(rows[0]))]
            body = rows
        return [dict(zip(keys, r)) for r in body]

    @staticmethod
    def write(path: Union[str, Path], rows: List[Dict[str, Any]],
              *, sheet_name: str = "Sheet1", headers: Optional[List[str]] = None) -> None:
        """把字典列表写入单个工作表。headers 省略时取第一行的键。"""
        _require()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        ExcelHelper._fill(ws, rows, headers)
        Path(path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)

    @staticmethod
    def write_sheets(path: Union[str, Path], sheets: Dict[str, List[Dict[str, Any]]]) -> None:
        """一次写入多个工作表：``{工作表名: 字典列表}``。"""
        _require()
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # 去掉默认空表
        for name, rows in sheets.items():
            ws = wb.create_sheet(title=name)
            ExcelHelper._fill(ws, rows)
        Path(path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)

    @staticmethod
    def _fill(ws, rows: List[Dict[str, Any]], headers: Optional[List[str]] = None) -> None:
        """把字典列表填入工作表，首行加粗表头。"""
        if not rows:
            return
        headers = headers or list(rows[0].keys())
        ws.append(headers)
        for cell in ws[1]:  # 表头加粗
            cell.font = openpyxl.styles.Font(bold=True)
        for row in rows:
            ws.append([row.get(h) for h in headers])


__all__ = ["ExcelHelper"]
